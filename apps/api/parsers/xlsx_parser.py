"""
XLSX/XLS Parser for the import pipeline.
Handles: header row detection (skips metadata rows), date conversion
(Excel serial → ISO), encoding, multiple sheets.

Uses openpyxl for .xlsx (read-only streaming mode).
Uses xlrd for .xls (legacy Excel).
"""

import re
import logging
import io
from typing import Optional

from parsers.base_parser import ParseResult, ColumnInfo, FileWarning

logger = logging.getLogger("import.xlsx_parser")

MAX_SAMPLE_VALUES = 5
MAX_HEADER_SCAN = 20


def _is_header_row(cells: list, min_fill_ratio: float = 0.4) -> bool:
    """Check if a row looks like a header (mostly non-empty strings, not numbers)."""
    non_empty = [c for c in cells if c is not None and str(c).strip()]
    if len(non_empty) < max(len(cells) * min_fill_ratio, 2):
        return False
    alpha_count = sum(1 for c in non_empty if isinstance(c, str) and re.search(r"[a-zA-Z]", c))
    return alpha_count >= len(non_empty) * 0.6


def _cell_to_str(value) -> str:
    """Convert a cell value to string, handling dates and numbers."""
    if value is None:
        return ""
    # openpyxl returns datetime objects for date-formatted cells
    from datetime import datetime, date
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _detect_date_format_from_cells(values: list) -> Optional[str]:
    """Detect date format from cell values."""
    from parsers.csv_parser import detect_date_format
    str_values = [_cell_to_str(v) for v in values if v is not None]
    return detect_date_format(str_values)


def _infer_domain_from_headers(headers: list[str]) -> Optional[str]:
    """Reuse CSV domain inference logic."""
    from parsers.csv_parser import infer_domain
    return infer_domain(headers)


def parse_xlsx(raw_data: bytes, filename: str = "unknown.xlsx") -> ParseResult:
    """
    Parse an XLSX file from raw bytes.
    Uses openpyxl in read-only mode for memory efficiency.
    """
    warnings: list[FileWarning] = []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return ParseResult(
            filename=filename,
            encoding_detected="utf-8",
            delimiter_detected=None,
            header_row=0,
            row_count=0,
            warnings=[FileWarning(field=None, message="openpyxl not installed", severity="red")],
        )

    try:
        wb = load_workbook(io.BytesIO(raw_data), read_only=True, data_only=True)
    except Exception as e:
        return ParseResult(
            filename=filename,
            encoding_detected="utf-8",
            delimiter_detected=None,
            header_row=0,
            row_count=0,
            warnings=[FileWarning(field=None, message=f"Cannot open XLSX: {e}", severity="red")],
        )

    # Use the first sheet
    ws = wb[wb.sheetnames[0]]

    # Read all rows (streaming)
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
        if len(all_rows) > 50000:
            warnings.append(FileWarning(field=None, message="File exceeds 50,000 rows — truncated", severity="amber"))
            break

    wb.close()

    if not all_rows:
        return ParseResult(
            filename=filename, encoding_detected="utf-8", delimiter_detected=None,
            header_row=0, row_count=0,
            warnings=[FileWarning(field=None, message="Spreadsheet is empty", severity="red")],
        )

    # Find header row (skip metadata rows — Sealogical pattern)
    header_idx = 0
    for i, row in enumerate(all_rows[:MAX_HEADER_SCAN]):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx > 0:
        warnings.append(FileWarning(
            field=None,
            message=f"Skipped {header_idx} metadata row(s) before header",
            severity="info",
            row=header_idx,
        ))

    headers = [_cell_to_str(c) for c in all_rows[header_idx]]
    # Remove trailing empty headers
    while headers and not headers[-1]:
        headers.pop()

    data_rows = all_rows[header_idx + 1:]

    # Build column info
    columns: list[ColumnInfo] = []
    all_date_values = []
    for col_idx, header in enumerate(headers):
        if not header:
            continue
        samples = []
        for row in data_rows[:MAX_SAMPLE_VALUES]:
            if col_idx < len(row) and row[col_idx] is not None:
                val = _cell_to_str(row[col_idx])
                if val:
                    samples.append(val)
                    # Collect date-like values
                    if re.search(r"\d{2,4}[/\-.]", val):
                        all_date_values.append(val)
        columns.append(ColumnInfo(source_name=header, sample_values=samples))

    # Detect date format
    date_format = _detect_date_format_from_cells(all_date_values) if all_date_values else None
    if date_format == "ambiguous":
        warnings.append(FileWarning(
            field=None,
            message="Date format is ambiguous (DD/MM vs MM/DD). User must confirm during mapping.",
            severity="amber",
        ))

    # Convert data rows to dicts
    row_dicts = []
    for row in data_rows:
        if not any(_cell_to_str(c) for c in row):
            continue
        row_dict = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = _cell_to_str(row[col_idx]) if col_idx < len(row) else ""
            row_dict[header] = value
        row_dicts.append(row_dict)

    # Infer domain
    domain = _infer_domain_from_headers(headers)

    logger.info(f"Parsed {filename}: XLSX, {len(row_dicts)} rows, {len(columns)} columns, domain={domain}")

    return ParseResult(
        filename=filename,
        encoding_detected="utf-8",
        delimiter_detected=None,
        header_row=header_idx,
        row_count=len(row_dicts),
        columns=columns,
        rows=row_dicts,
        date_format_detected=date_format,
        warnings=warnings,
        domain_hint=domain,
    )


def parse_xls(raw_data: bytes, filename: str = "unknown.xls") -> ParseResult:
    """
    Parse a legacy .xls file using xlrd.
    """
    warnings: list[FileWarning] = []

    try:
        import xlrd
    except ImportError:
        return ParseResult(
            filename=filename, encoding_detected="utf-8", delimiter_detected=None,
            header_row=0, row_count=0,
            warnings=[FileWarning(field=None, message="xlrd not installed", severity="red")],
        )

    try:
        wb = xlrd.open_workbook(file_contents=raw_data)
    except Exception as e:
        return ParseResult(
            filename=filename, encoding_detected="utf-8", delimiter_detected=None,
            header_row=0, row_count=0,
            warnings=[FileWarning(field=None, message=f"Cannot open XLS: {e}", severity="red")],
        )

    sheet = wb.sheet_by_index(0)

    # Read all rows
    all_rows = []
    for row_idx in range(min(sheet.nrows, 50001)):
        row_values = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row_values.append(dt.strftime("%Y-%m-%d"))
                except Exception:
                    row_values.append(str(cell.value))
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                # Integer check
                if cell.value == int(cell.value):
                    row_values.append(str(int(cell.value)))
                else:
                    row_values.append(str(cell.value))
            else:
                row_values.append(str(cell.value).strip() if cell.value else "")
        all_rows.append(row_values)

    if not all_rows:
        return ParseResult(
            filename=filename, encoding_detected="utf-8", delimiter_detected=None,
            header_row=0, row_count=0,
            warnings=[FileWarning(field=None, message="Spreadsheet is empty", severity="red")],
        )

    # Find header row
    header_idx = 0
    for i, row in enumerate(all_rows[:MAX_HEADER_SCAN]):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx > 0:
        warnings.append(FileWarning(
            field=None, message=f"Skipped {header_idx} metadata row(s) before header",
            severity="info", row=header_idx,
        ))

    headers = [str(c).strip() for c in all_rows[header_idx]]
    while headers and not headers[-1]:
        headers.pop()

    data_rows = all_rows[header_idx + 1:]

    columns = []
    for col_idx, header in enumerate(headers):
        if not header:
            continue
        samples = []
        for row in data_rows[:MAX_SAMPLE_VALUES]:
            if col_idx < len(row) and row[col_idx]:
                samples.append(str(row[col_idx]).strip())
        columns.append(ColumnInfo(source_name=header, sample_values=samples))

    row_dicts = []
    for row in data_rows:
        if not any(str(c).strip() for c in row):
            continue
        row_dict = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = str(row[col_idx]).strip() if col_idx < len(row) else ""
            row_dict[header] = value
        row_dicts.append(row_dict)

    domain = _infer_domain_from_headers(headers)

    logger.info(f"Parsed {filename}: XLS, {len(row_dicts)} rows, {len(columns)} columns, domain={domain}")

    return ParseResult(
        filename=filename, encoding_detected="utf-8", delimiter_detected=None,
        header_row=header_idx, row_count=len(row_dicts), columns=columns,
        rows=row_dicts, date_format_detected=None, warnings=warnings, domain_hint=domain,
    )
