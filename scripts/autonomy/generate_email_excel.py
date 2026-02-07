#!/usr/bin/env python3
"""
Generate Excel Sheet with 100 Test Emails

Creates realistic email content for manual sending via workflow.
Output: test_emails_100.xlsx

Columns:
- Subject
- Body
- From
- To
- Scenario
- Expected_Object_Type
- Expected_Object_ID
- Attachments_List
- Attachment_Notes
"""

import os
import sys
import json
import random
from datetime import datetime
from typing import List, Dict, Any

# Add parent to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../apps/api'))

from integrations.supabase import get_supabase_client

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_sampled_data() -> Dict[str, Any]:
    """Load sampled real data."""
    sample_path = 'test-results/autonomy/sampled_data.json'

    if not os.path.exists(sample_path):
        print(f"⚠ No sampled data found. Sampling now...")
        # Sample data
        os.makedirs('test-results/autonomy', exist_ok=True)

        os.environ['SUPABASE_URL'] = os.getenv('yTEST_YACHT_001_SUPABASE_URL', '')
        os.environ['SUPABASE_SERVICE_KEY'] = os.getenv('yTEST_YACHT_001_SUPABASE_SERVICE_KEY', '')

        supabase = get_supabase_client()
        yacht_id = os.getenv('TEST_YACHT_ID', '00000000-0000-0000-0000-000000000001')

        # Sample parts
        parts_result = supabase.table('pms_parts').select(
            'id, name, part_number, manufacturer, category'
        ).eq('yacht_id', yacht_id).not_.is_('part_number', 'null').limit(20).execute()

        # Sample equipment
        equipment_result = supabase.table('equipment').select(
            'id, name, serial_number, model, manufacturer'
        ).eq('yacht_id', yacht_id).not_.is_('serial_number', 'null').limit(20).execute()

        # Sample work orders
        wo_result = supabase.table('pms_work_orders').select(
            'id, wo_number, title, status, equipment_id'
        ).eq('yacht_id', yacht_id).order('created_at', desc=True).limit(20).execute()

        # Sample vendors (if table exists)
        try:
            vendors_result = supabase.table('vendors').select(
                'id, name, email, domain, category'
            ).eq('yacht_id', yacht_id).not_.is_('email', 'null').limit(20).execute()
        except:
            vendors_result = type('obj', (object,), {'data': []})()

        sampled_data = {
            'yacht_id': yacht_id,
            'parts': parts_result.data or [],
            'equipment': equipment_result.data or [],
            'work_orders': wo_result.data or [],
            'vendors': vendors_result.data or [],
        }

        with open(sample_path, 'w') as f:
            json.dump(sampled_data, f, indent=2, default=str)

        print(f"✓ Sampled data saved to {sample_path}")
        return sampled_data

    with open(sample_path, 'r') as f:
        return json.load(f)


def generate_wo_explicit_email(wo: Dict, idx: int) -> Dict[str, Any]:
    """Generate explicit WO ID email (L1)."""
    wo_number = wo['wo_number']

    subjects = [
        f"Re: WO-{wo_number} - Status Update",
        f"Work Order {wo_number} - Parts Arrived",
        f"{wo_number}: Completion Estimate",
        f"FW: WO-{wo_number} Vendor Quote",
        f"Update on {wo_number}",
        f"WO-{wo_number} - Final Invoice Attached",
    ]

    subject = random.choice(subjects)

    bodies = [
        f"""Hi Team,

Just wanted to update you on work order {wo_number} ({wo['title']}).

Current status: {wo['status']}

Parts have been ordered and should arrive by end of week. Will schedule installation once they're in.

Let me know if you need anything else.

Thanks,
Alex""",
        f"""Hello,

Regarding WO-{wo_number}, we've completed the initial inspection.

Work Order: {wo['title']}
Status: {wo['status']}

Estimated completion: 5-7 business days
Parts on order: Yes

Will keep you posted.

Best,
Service Team""",
        f"""Quick update on {wo_number}:

- Technician assigned
- Parts sourced
- Target completion: Next week

Work order title: {wo['title']}

Thanks,
Alex""",
    ]

    body = random.choice(bodies)

    # Attachments
    attachments = []
    if random.random() < 0.4:  # 40% chance of invoice
        attachments.append({
            'name': f'Invoice_WO{wo_number}_{idx}.pdf',
            'type': 'invoice',
            'notes': f'Contains WO-{wo_number} in header and line items'
        })

    if random.random() < 0.3:  # 30% chance of quote
        attachments.append({
            'name': f'Quote_{wo_number}_{idx}.pdf',
            'type': 'quote',
            'notes': f'Vendor quote referencing WO-{wo_number}'
        })

    return {
        'subject': subject,
        'body': body,
        'scenario': 'wo_explicit',
        'expected_object_type': 'work_order',
        'expected_object_id': wo['id'],
        'attachments': attachments,
    }


def generate_part_number_email(part: Dict, equipment_list: List[Dict], idx: int) -> Dict[str, Any]:
    """Generate part number email (L2.5)."""
    part_number = part['part_number']
    part_name = part['name']
    manufacturer = part.get('manufacturer', 'OEM')

    equipment = random.choice(equipment_list) if equipment_list else None
    equipment_name = equipment['name'] if equipment else 'the equipment'

    subjects = [
        f"Part {part_number} - Availability Question",
        f"Ordering {part_number} for {equipment_name}",
        f"Quote Request: {part_number}",
        f"{part_number} - Installation Manual Needed",
        f"P/N: {part_number} - Stock Status",
        f"Part Number {part_number} ({manufacturer})",
    ]

    subject = random.choice(subjects)

    bodies = [
        f"""Hi,

We need part number {part_number} ({part_name}) for {equipment_name}.

Manufacturer: {manufacturer}
Quantity needed: 2

Can you provide pricing and lead time?

Also, do you have the installation manual for this part?

Thanks,
Alex""",
        f"""Hello,

Checking availability on part {part_number}.

Description: {part_name}
Application: {equipment_name}
Manufacturer: {manufacturer}

Please send quote when you have a chance.

Best,
Service Team""",
        f"""Quick question about P/N {part_number}:

We're working on {equipment_name} and need this part urgently.

Part: {part_name}
Mfg: {manufacturer}

Do you have stock? What's the price?

Thanks,
Alex""",
    ]

    body = random.choice(bodies)

    # Attachments
    attachments = []
    if random.random() < 0.5:  # 50% chance of datasheet
        attachments.append({
            'name': f'Datasheet_{part_number.replace("/", "_")}_{idx}.pdf',
            'type': 'datasheet',
            'notes': f'Technical specs for {part_number}, mentions part number in title and body'
        })

    if random.random() < 0.3:  # 30% chance of install manual
        attachments.append({
            'name': f'Install_Manual_{part_number.replace("/", "_")}_{idx}.pdf',
            'type': 'manual',
            'notes': f'Installation instructions for {part_number}'
        })

    return {
        'subject': subject,
        'body': body,
        'scenario': 'part_number',
        'expected_object_type': 'part',
        'expected_object_id': part['id'],
        'attachments': attachments,
    }


def generate_equipment_serial_email(equipment: Dict, idx: int) -> Dict[str, Any]:
    """Generate equipment serial number email (L2.5)."""
    serial = equipment['serial_number']
    name = equipment['name']
    model = equipment.get('model', 'N/A')
    manufacturer = equipment.get('manufacturer', 'OEM')

    subjects = [
        f"Service Request for {name} (S/N: {serial})",
        f"{name} Maintenance - Serial {serial}",
        f"Re: {name} ({model}) - Troubleshooting",
        f"{manufacturer} {model} - Service History Request",
        f"Equipment Issue: {name} S/N {serial}",
    ]

    subject = random.choice(subjects)

    bodies = [
        f"""Hi Team,

We're having an issue with our {name}.

Serial Number: {serial}
Model: {model}
Manufacturer: {manufacturer}

The equipment is showing error codes and needs service. Can someone come take a look?

Thanks,
Alex""",
        f"""Hello,

Requesting maintenance on {name} (S/N: {serial}).

Equipment Details:
- Model: {model}
- Manufacturer: {manufacturer}
- Serial: {serial}

Last service was 6 months ago. Due for routine maintenance.

Please advise on scheduling.

Best,
Service Team""",
        f"""Quick question about {name} with serial number {serial}:

We need the service history and any available documentation for this unit.

Model: {model}
Mfg: {manufacturer}

Thanks,
Alex""",
    ]

    body = random.choice(bodies)

    # Attachments
    attachments = []
    if random.random() < 0.4:  # 40% chance of service report
        attachments.append({
            'name': f'Service_Report_{serial}_{idx}.pdf',
            'type': 'service_report',
            'notes': f'Service report for {name} S/N {serial}, contains serial in header'
        })

    if random.random() < 0.3:  # 30% chance of error log
        attachments.append({
            'name': f'Error_Log_{serial}_{idx}.txt',
            'type': 'log',
            'notes': f'System logs from {name}, serial {serial} in metadata'
        })

    return {
        'subject': subject,
        'body': body,
        'scenario': 'equipment_serial',
        'expected_object_type': 'equipment',
        'expected_object_id': equipment['id'],
        'attachments': attachments,
    }


def generate_warranty_claim_email(equipment: Dict, idx: int) -> Dict[str, Any]:
    """Generate warranty claim email (L2.5)."""
    serial = equipment['serial_number']
    name = equipment['name']
    model = equipment.get('model', 'N/A')
    manufacturer = equipment.get('manufacturer', 'OEM')

    subjects = [
        f"Warranty Claim: {name} Failure",
        f"Re: Warranty Status for {name}",
        f"Claim Documentation for {name} (S/N: {serial})",
        f"Warranty Service Request - {manufacturer} {model}",
        f"Equipment Warranty Claim - {name}",
    ]

    subject = random.choice(subjects)

    bodies = [
        f"""Hello Warranty Team,

We need to file a warranty claim for {name}.

Equipment Details:
- Serial Number: {serial}
- Model: {model}
- Manufacturer: {manufacturer}

Issue: Complete equipment failure, will not start.

Purchase date was within warranty period. Please advise on next steps and documentation needed.

Thanks,
Alex""",
        f"""Hi,

Warranty claim request for {name} (S/N: {serial}).

The equipment failed unexpectedly and appears to be a manufacturing defect.

Model: {model}
Mfg: {manufacturer}

Can you review warranty coverage and let us know the claim process?

Attached: Photos of failure, original invoice

Best,
Service Team""",
        f"""Warranty inquiry for {name}:

Our {manufacturer} {model} unit (serial {serial}) has failed.

We believe this is covered under warranty. Can you confirm coverage and provide RMA instructions?

Thanks,
Alex""",
    ]

    body = random.choice(bodies)

    # Attachments
    attachments = []
    if random.random() < 0.7:  # 70% chance of photos
        attachments.append({
            'name': f'Failure_Photos_{serial}_{idx}.zip',
            'type': 'photos',
            'notes': f'Photos of {name} failure, serial {serial} visible in images'
        })

    if random.random() < 0.6:  # 60% chance of original invoice
        attachments.append({
            'name': f'Original_Invoice_{serial}_{idx}.pdf',
            'type': 'invoice',
            'notes': f'Purchase invoice showing {name}, S/N {serial}, and warranty terms'
        })

    if random.random() < 0.4:  # 40% chance of warranty doc
        attachments.append({
            'name': f'Warranty_Certificate_{idx}.pdf',
            'type': 'warranty',
            'notes': f'Original warranty certificate for {name}'
        })

    return {
        'subject': subject,
        'body': body,
        'scenario': 'warranty_claim',
        'expected_object_type': 'equipment',
        'expected_object_id': equipment['id'],
        'attachments': attachments,
    }


def generate_vendor_generic_email(vendor: Dict, idx: int) -> Dict[str, Any]:
    """Generate generic vendor email (L4)."""
    vendor_name = vendor['name']
    vendor_email = vendor['email']

    subjects = [
        f"Quote Request for Upcoming Service",
        f"Re: Parts Pricing Discussion",
        f"Follow-up: Service Proposal",
        f"Availability Check",
        f"General Inquiry - {vendor_name}",
        f"Service Options for Marine Equipment",
    ]

    subject = random.choice(subjects)

    bodies = [
        f"""Hi,

This is Alex from the yacht maintenance team. We're looking for service options for upcoming maintenance work.

Can you provide a general quote for your services?

Thanks,
Alex""",
        f"""Hello,

Following up on our previous conversation about parts and service.

We'd like to get pricing on general maintenance supplies.

Let me know when you have availability to discuss.

Best,
Service Team""",
        f"""Hi {vendor_name},

Quick inquiry about your service offerings. We may need some work done in the next few weeks.

Can you send over your standard pricing and availability?

Thanks,
Alex""",
    ]

    body = random.choice(bodies)

    # Attachments
    attachments = []
    if random.random() < 0.3:  # 30% chance of specs
        attachments.append({
            'name': f'Equipment_Specs_{idx}.pdf',
            'type': 'specs',
            'notes': f'General equipment specifications for quote request'
        })

    return {
        'subject': subject,
        'body': body,
        'from_email': vendor_email,  # Send FROM vendor
        'scenario': 'vendor_generic',
        'expected_object_type': 'vendor',
        'expected_object_id': vendor['id'],
        'attachments': attachments,
    }


def main():
    print("=" * 60)
    print("Generating 100 Test Emails for Excel")
    print("=" * 60)
    print()

    # Load sampled data
    sampled_data = load_sampled_data()

    parts = sampled_data.get('parts', [])
    equipment = sampled_data.get('equipment', [])
    work_orders = sampled_data.get('work_orders', [])
    vendors = sampled_data.get('vendors', [])

    print(f"Sampled data: {len(parts)} parts, {len(equipment)} equipment, {len(work_orders)} WOs, {len(vendors)} vendors")
    print()

    # Generate email distribution
    emails = []

    # 20 explicit WO emails (L1 baseline)
    for i in range(20):
        if work_orders:
            wo = random.choice(work_orders)
            email = generate_wo_explicit_email(wo, i)
            email['index'] = len(emails) + 1
            email['from_email'] = 'x@alex-short.com'
            email['to_email'] = 'x@alex-short.com'
            emails.append(email)

    # 30 part number emails (L2.5)
    for i in range(30):
        if parts:
            part = random.choice(parts)
            email = generate_part_number_email(part, equipment, i)
            email['index'] = len(emails) + 1
            email['from_email'] = 'x@alex-short.com'
            email['to_email'] = 'x@alex-short.com'
            emails.append(email)

    # 25 equipment serial emails (L2.5)
    for i in range(25):
        if equipment:
            eq = random.choice(equipment)
            email = generate_equipment_serial_email(eq, i)
            email['index'] = len(emails) + 1
            email['from_email'] = 'x@alex-short.com'
            email['to_email'] = 'x@alex-short.com'
            emails.append(email)

    # 15 warranty claim emails (L2.5)
    for i in range(15):
        if equipment:
            eq = random.choice(equipment)
            email = generate_warranty_claim_email(eq, i)
            email['index'] = len(emails) + 1
            email['from_email'] = 'x@alex-short.com'
            email['to_email'] = 'x@alex-short.com'
            emails.append(email)

    # 10 vendor generic emails (L4)
    for i in range(10):
        if vendors:
            vendor = random.choice(vendors)
            email = generate_vendor_generic_email(vendor, i)
            email['index'] = len(emails) + 1
            email['to_email'] = 'x@alex-short.com'
            if 'from_email' not in email:
                email['from_email'] = 'x@alex-short.com'
            emails.append(email)

    # Shuffle for randomness
    random.shuffle(emails)

    # Re-index
    for i, email in enumerate(emails):
        email['index'] = i + 1

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Emails"

    # Headers
    headers = [
        'Index',
        'Subject',
        'Body',
        'From',
        'To',
        'Scenario',
        'Expected_Object_Type',
        'Expected_Object_ID',
        'Attachments_Count',
        'Attachment_1_Name',
        'Attachment_1_Type',
        'Attachment_1_Notes',
        'Attachment_2_Name',
        'Attachment_2_Type',
        'Attachment_2_Notes',
        'Attachment_3_Name',
        'Attachment_3_Type',
        'Attachment_3_Notes',
    ]

    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data
    for email in emails:
        attachments = email.get('attachments', [])

        row = [
            email['index'],
            email['subject'],
            email['body'],
            email.get('from_email', 'x@alex-short.com'),
            email.get('to_email', 'x@alex-short.com'),
            email['scenario'],
            email.get('expected_object_type', ''),
            email.get('expected_object_id', ''),
            len(attachments),
        ]

        # Add up to 3 attachments
        for i in range(3):
            if i < len(attachments):
                att = attachments[i]
                row.extend([
                    att['name'],
                    att['type'],
                    att['notes']
                ])
            else:
                row.extend(['', '', ''])

        ws.append(row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 38
    ws.column_dimensions['I'].width = 12

    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws.column_dimensions[col].width = 25

    # Wrap text
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save
    output_file = 'test-results/autonomy/test_emails_100.xlsx'
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)

    print(f"✓ Generated {len(emails)} test emails")
    print(f"✓ Saved to: {output_file}")
    print()

    # Print summary
    print("Distribution:")
    scenario_counts = {}
    for email in emails:
        scenario = email['scenario']
        scenario_counts[scenario] = scenario_counts.get(scenario, 0) + 1

    for scenario, count in sorted(scenario_counts.items()):
        print(f"  {scenario}: {count}")

    print()
    print("=" * 60)
    print("Next Steps:")
    print("1. Open: test-results/autonomy/test_emails_100.xlsx")
    print("2. Use your workflow to send these emails")
    print("3. Emails will flow through: Graph API → Email Watcher → Token Extraction → Link Suggester")
    print("4. Run validation: python scripts/autonomy/validate_autolinking.py")
    print("=" * 60)


if __name__ == '__main__':
    main()
